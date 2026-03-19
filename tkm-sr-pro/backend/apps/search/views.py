from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
import os
import json
import openai
import requests as http_requests
import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from django.utils import timezone
from .utils.api_clients import SearchManager
from .utils.mesh_expander import MeSHExpander
from .utils.tkm_expander import TKMExpander
from .serializers.search_serializers import SearchQuerySerializer, LiteratureRecordSerializer, ExpandQuerySerializer
from common.utils.deduplicator import HybridDeduplicator
from .models import LiteratureRecord, SearchProject

import re
import sys

class TestView(APIView):
    def get(self, request):
        return Response({"status": "ok"}, status=status.HTTP_200_OK)

def perform_structured_expansion(disease_input, formula_input, category, include_rct, api_key):
    print(f"DEBUG: START EXPAND - A='{disease_input}', B='{formula_input}', C='{category}'")
    sys.stdout.flush()
    
    tkm_expander = TKMExpander()
    mesh_expander = MeSHExpander()
    
    d_parts = [] 
    f_parts = []

    # Process A (Population/Disease)
    a_terms = re.findall(r'\'[^\']+\'|"[^"]+"|\S+', disease_input)
    for t in a_terms:
        clean_t = t.strip("'").strip('"')
        if not clean_t: continue
        if re.search(r'[a-zA-Z]', clean_t):
            try:
                expanded = mesh_expander.expand_query(clean_t)
                d_parts.append(expanded)
            except:
                d_parts.append(f'"{clean_t}"' if " " in clean_t else clean_t)
        else:
            d_parts.append(f'"{clean_t}"' if " " in clean_t else clean_t)

    # Process B (Intervention/Formula)
    b_terms = re.findall(r'\'[^\']+\'|"[^"]+"|\S+', formula_input)
    for t in b_terms:
        clean_t = t.strip("'").strip('"')
        if not clean_t: continue
        # If in local dictionary, use synonyms
        if clean_t in tkm_expander.tkm_prescriptions:
            syns = tkm_expander.tkm_prescriptions[clean_t]
            f_parts.append(f'("{clean_t}" OR ' + " OR ".join(f'"{s}"' if " " in s else s for s in syns) + ')')
        else:
            # Add as is
            f_parts.append(f'"{clean_t}"' if " " in clean_t else clean_t)

    # Phase 2: GPT Enrichment (Optional)
    if api_key:
        prompt = f"Expert Medical Librarian: Expand A: {disease_input} and B: {formula_input} into English/CJK synonyms. Return JSON {{'expanded_disease': '(t1 OR t2)', 'expanded_formula': '(t3 OR t4)'}}"
        try:
            headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
            data = {"model": "gpt-4o", "messages": [{"role": "user", "content": prompt}], "response_format": {"type": "json_object"}, "temperature": 0.1}
            response = http_requests.post("https://api.openai.com/v1/chat/completions", headers=headers, json=data, timeout=30)
            if response.status_code == 200:
                parsed = json.loads(response.json()["choices"][0]["message"]["content"])
                gpt_d = parsed.get("expanded_disease", "")
                gpt_f = parsed.get("expanded_formula", "")
                if gpt_d: d_parts.append(gpt_d)
                if gpt_f: f_parts.append(gpt_f)
        except Exception as e:
            print(f"DEBUG: GPT ERROR: {e}")

    # Final Construction: (A) AND ((B) OR (C)) AND (D)
    final_groups = []
    
    # Section A
    if d_parts:
        a_inner = " AND ".join(sorted(list(set(d_parts))))
        final_groups.append(f"({a_inner})")
    
    # Section B + C
    lookup_cat = category if category else "전체"
    cat_mesh = tkm_expander.tkm_categories.get(lookup_cat, "")
    b_formatted = ""
    if f_parts:
        b_inner = " OR ".join(sorted(list(set(f_parts))))
        b_formatted = f"({b_inner})"
    
    if b_formatted and cat_mesh:
        final_groups.append(f"(({b_formatted}) OR ({cat_mesh}))")
    elif b_formatted:
        final_groups.append(f"({b_formatted})")
    elif cat_mesh:
        final_groups.append(f"({cat_mesh})")
        
    # Section D
    if include_rct:
        rct_str = "'randomized controlled trial'/exp OR 'controlled clinical trial' OR random* OR 'placebo' OR trial"
        final_groups.append(f"({rct_str})")
        
    res_query = " AND ".join(final_groups)
    print(f"DEBUG: FINAL EXPANSION (Strict): {res_query}")
    sys.stdout.flush()
    return res_query

class ExpandQueryView(APIView):

    def post(self, request):
        print(f"DEBUG: ExpandQueryView index data: {request.data}")
        sys.stdout.flush()
        serializer = ExpandQuerySerializer(data=request.data)
        if serializer.is_valid():
            disease = serializer.validated_data.get('disease', '').strip()
            formula = serializer.validated_data.get('formula', '').strip()
            category = serializer.validated_data.get('category', '')
            include_rct = serializer.validated_data.get('include_rct', True)
            api_key = serializer.validated_data.get('api_key', '')
            
            final_query = perform_structured_expansion(disease, formula, category, include_rct, api_key)
            return Response({"expanded_query": final_query}, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class FederatedSearchView(APIView):
    def post(self, request):
        print(f"DEBUG: FederatedSearchView index data: {request.data}")
        sys.stdout.flush()
        serializer = SearchQuerySerializer(data=request.data)
        if serializer.is_valid():
            query = serializer.validated_data.get('query', '')
            disease = serializer.validated_data.get('disease', '').strip()
            formula = serializer.validated_data.get('formula', '').strip()
            category = serializer.validated_data.get('category', '전체')
            include_rct = serializer.validated_data.get('include_rct', True)
            exact_query = serializer.validated_data.get('exact_query', '')
            api_key = serializer.validated_data.get('api_key', '')
            max_results = serializer.validated_data.get('max_results', 300)
            dbs = serializer.validated_data.get('dbs', ['PubMed', 'Cochrane', 'ScienceON', 'RISS', 'CiNii'])
            
            # 1. Determine Expanded Query (Mainly for PubMed)
            if exact_query:
                pubmed_query = exact_query
            else:
                if disease or formula or (category and category != "전체" and category != ""):
                    pubmed_query = perform_structured_expansion(disease, formula, category, include_rct, api_key)
                else:
                    expander = MeSHExpander()
                    pubmed_query = expander.expand_query(query)
            
            # 2. Determine Plain Query (For ScienceON, RISS, CiNii)
            # Remove PubMed-specific tags like [MeSH] or [Title/Abstract]
            # And potentially remove RCT filter if it's too restrictive for non-English DBs
            plain_query = re.sub(r'\[[^\]]+\]', '', pubmed_query)
            # If the user customized the textarea, we use the cleaned textarea as the plain query
            # Otherwise we use the concatenated disease+formula inputs
            if not exact_query and (disease or formula):
                # Clean version of concatenated inputs is often safer for simplistic DBs
                non_pubmed_query = f"{disease} {formula}".strip()
            else:
                # Use cleaned expanded query
                non_pubmed_query = plain_query

            # Ensure we have a query for ScienceON/RISS/etc
            if not non_pubmed_query:
                non_pubmed_query = query if query else "한의학"

            manager = SearchManager()
            # We pass both: pubmed_query (with MeSH) and non_pubmed_query (clean structure)
            # We use 'query' parameter of manager to hold the cleaned one, and 'expanded_query' for PubMed
            results = manager.federated_search(non_pubmed_query, expanded_query=pubmed_query, dbs=dbs, max_results=max_results)
            
            return Response({
                "query": non_pubmed_query,
                "expanded_query": pubmed_query,
                "results": results
            }, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class SaveRecordsView(APIView):
    def post(self, request):
        records_data = request.data
        if not records_data:
            return Response({"error": "No records provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Determine target project: use _project_id if provided, else find/create default
        raw_project_id = records_data[0].get('_project_id') if records_data else None
        if raw_project_id:
            try:
                target_project = SearchProject.objects.get(pk=raw_project_id)
            except SearchProject.DoesNotExist:
                target_project = None
        else:
            target_project = None

        if not target_project:
            target_project, _ = SearchProject.objects.get_or_create(
                name="Default SR Project",
                defaults={"description": "Automatically created project for search results."}
            )

        # Fields allowed by the LiteratureRecord model
        ALLOWED_FIELDS = {'title', 'abstract', 'authors', 'journal', 'year', 'doi', 'pmid', 'keywords', 'source_db', 'project'}

        cleaned = []
        for item in records_data:
            record = {k: v for k, v in item.items() if k in ALLOWED_FIELDS}
            record["project"] = target_project.id
            record["source_db"] = record.get("source_db") or item.get("source", "PubMed")
            # Skip records that have no title (required field)
            title = record.get("title", "")
            if not title or not str(title).strip():
                continue
            record["title"] = str(title).strip()
            cleaned.append(record)

        if not cleaned:
            return Response({"error": "No valid records to save (all missing title)"}, status=status.HTTP_400_BAD_REQUEST)

        serializer = LiteratureRecordSerializer(data=cleaned, many=True)
        if serializer.is_valid():
            records = serializer.save()
            return Response({"saved_count": len(records), "skipped": len(records_data) - len(cleaned)}, status=status.HTTP_201_CREATED)
        return Response({"error": "Validation failed", "detail": serializer.errors[:3]}, status=status.HTTP_400_BAD_REQUEST)


class DeduplicateRecordsView(APIView):
    def post(self, request):
        use_db = request.data.get('use_db', False)
        records = request.data.get('records', [])
        project_id = request.data.get('project_id')

        if use_db and not records:
            qs = LiteratureRecord.objects.filter(status=LiteratureRecord.Status.IMPORTED)
            if project_id:
                qs = qs.filter(project_id=project_id)
            records = list(qs.values('id', 'title', 'authors', 'abstract', 'year', 'pmid', 'doi'))

        if not records:
            return Response({"error": "No records provided or found in DB"}, status=status.HTTP_400_BAD_REQUEST)

        deduplicator = HybridDeduplicator()
        results = deduplicator.deduplicate(records)

        if use_db:
            duplicate_ids = [res['record_b_id'] for res in results if res['status'] == 'auto']
            review_ids = [res['record_b_id'] for res in results if res['status'] == 'review']
            if duplicate_ids:
                LiteratureRecord.objects.filter(id__in=duplicate_ids).update(status=LiteratureRecord.Status.DEDUP_REJECTED)
            if review_ids:
                LiteratureRecord.objects.filter(id__in=review_ids).update(status=LiteratureRecord.Status.DEDUP_REVIEW)
            all_ids = [r['id'] for r in records]
            remaining_ids = set(all_ids) - set(duplicate_ids) - set(review_ids)
            if remaining_ids:
                LiteratureRecord.objects.filter(id__in=remaining_ids).update(status=LiteratureRecord.Status.SCREENING_PENDING)

        return Response({
            "results": results,
            "processed_count": len(records),
            "duplicates_found": len(results)
        }, status=status.HTTP_200_OK)

class ImportedRecordsView(APIView):
    def get(self, request):
        project_id = request.query_params.get('project_id')
        qs = LiteratureRecord.objects.filter(status=LiteratureRecord.Status.IMPORTED)
        if project_id:
            qs = qs.filter(project_id=project_id)
        records = list(qs.values('id', 'title', 'authors', 'abstract', 'year', 'pmid', 'doi', 'source_db'))
        return Response({"records": records}, status=status.HTTP_200_OK)

class ScreeningPendingView(APIView):
    """Return records that are waiting for RCT screening."""
    def get(self, request):
        project_id = request.query_params.get('project_id')
        qs = LiteratureRecord.objects.filter(status=LiteratureRecord.Status.SCREENING_PENDING)
        if project_id:
            qs = qs.filter(project_id=project_id)
        records = list(qs.values('id', 'title', 'abstract', 'authors', 'year', 'pmid', 'source_db'))
        return Response({"records": records, "count": len(records)}, status=status.HTTP_200_OK)

class RctPredictView(APIView):
    """Proxy a single record to the AI engine for RCT prediction."""
    def post(self, request):
        ai_url = os.getenv('AI_ENGINE_URL', 'http://ai_engine:8001')
        payload = {
            "title": request.data.get('title', ''),
            "abstract": request.data.get('abstract', ''),
            "keywords": request.data.get('keywords', ''),
        }
        try:
            resp = http_requests.post(f"{ai_url}/api/v1/ai/predict_rct", json=payload, timeout=15)
            resp.raise_for_status()
            return Response(resp.json(), status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

class RctDecisionView(APIView):
    """Save human final decision on a record (include/exclude)."""
    def post(self, request):
        record_id = request.data.get('record_id')
        decision = request.data.get('decision')  # 'include' or 'exclude'
        if not record_id or decision not in ('include', 'exclude'):
            return Response({"error": "Invalid parameters"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            rec = LiteratureRecord.objects.get(id=record_id)
            if decision == 'include':
                rec.status = LiteratureRecord.Status.RCT_INCLUDED
            else:
                rec.status = LiteratureRecord.Status.RCT_EXCLUDED
            rec.save()
            return Response({"ok": True, "id": record_id, "status": rec.status}, status=status.HTTP_200_OK)
        except LiteratureRecord.DoesNotExist:
            return Response({"error": "Record not found"}, status=status.HTTP_404_NOT_FOUND)

class DashboardStatsView(APIView):
    def get(self, request):
        project_id = request.query_params.get('project_id')
        try:
            qs = LiteratureRecord.objects
            if project_id:
                qs = qs.filter(project_id=project_id)

            total = qs.count()
            if total == 0:
                return Response({
                    "totalSearched": 0,
                    "deduplicated": 0,
                    "reviewNeeded": 0,
                    "pendingScreening": 0,
                    "rctFiltered": 0,
                    "extracted": 0
                }, status=status.HTTP_200_OK)

            dedup_rejected = qs.filter(status=LiteratureRecord.Status.DEDUP_REJECTED).count()
            review_needed = qs.filter(status=LiteratureRecord.Status.DEDUP_REVIEW).count()
            pending_screening = qs.filter(status=LiteratureRecord.Status.SCREENING_PENDING).count()
            dedup = total - dedup_rejected - review_needed

            rct = qs.filter(
                status__in=[LiteratureRecord.Status.RCT_INCLUDED,
                             LiteratureRecord.Status.FULLTEXT_INCLUDED,
                             LiteratureRecord.Status.FULLTEXT_EXCLUDED,
                             LiteratureRecord.Status.EXTRACTED]
            ).count()

            extracted = qs.filter(status=LiteratureRecord.Status.EXTRACTED).count()
            
            return Response({
                "totalSearched": total,
                "deduplicated": dedup,
                "reviewNeeded": review_needed,
                "pendingScreening": pending_screening,
                "rctFiltered": rct,
                "extracted": extracted
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            print(f"Stats error: {e}")
            return Response({
                "totalSearched": 0,
                "deduplicated": 0,
                "reviewNeeded": 0,
                "pendingScreening": 0,
                "rctFiltered": 0,
                "extracted": 0
            }, status=status.HTTP_200_OK)

class RctIncludedListView(APIView):
    """Return records that have been confirmed as RCT_INCLUDED or FULLTEXT_INCLUDED, ready for extraction."""
    def get(self, request):
        project_id = request.query_params.get('project_id')
        qs = LiteratureRecord.objects.filter(
            status__in=[
                LiteratureRecord.Status.RCT_INCLUDED,
                LiteratureRecord.Status.FULLTEXT_INCLUDED,
                LiteratureRecord.Status.EXTRACTED
            ]
        )
        if project_id:
            qs = qs.filter(project_id=project_id)
        records = list(qs.values('id', 'title', 'abstract', 'authors', 'year', 'pmid', 'doi', 'source_db', 'status', 'full_text', 'pico_data', 'pico_last_extracted_at', 'pico_confirmed_at'))
        return Response({"records": records, "count": len(records)}, status=status.HTTP_200_OK)

class PicoExtractView(APIView):
    """
    Proxy abstract text to AI engine for PICO extraction, then save to DB.
    """
    def post(self, request):
        record_id = request.data.get('record_id')
        title = request.data.get('title', '')
        abstract = request.data.get('abstract', '')
        full_text = request.data.get('full_text', '')

        ai_url = os.getenv('AI_ENGINE_URL', 'http://ai_engine:8001')
        payload = {"title": title, "abstract": abstract, "full_text": full_text}

        try:
            resp = http_requests.post(f"{ai_url}/api/v1/ai/extract_pico", json=payload, timeout=30)
            resp.raise_for_status()
            pico_data = resp.json()

            # If record_id provided, update status to EXTRACTED in DB
            if record_id:
                try:
                    rec = LiteratureRecord.objects.get(id=record_id)
                    rec.pico_data = pico_data
                    rec.pico_last_extracted_at = timezone.now()
                    rec.save()
                    pico_data['record_id'] = record_id
                    pico_data['saved_to_db'] = True
                    pico_data['pico_last_extracted_at'] = rec.pico_last_extracted_at
                except LiteratureRecord.DoesNotExist:
                    pico_data['saved_to_db'] = False

            return Response(pico_data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

class PicoExtractGptView(APIView):
    """
    Enhanced PICO extraction using ChatGPT (GPT-4o) with user-provided API key.
    """
    def post(self, request):
        api_key = request.data.get('api_key')
        record_id = request.data.get('record_id')
        title = request.data.get('title', '')
        abstract = request.data.get('abstract', '')
        full_text = request.data.get('full_text', '')

        if not api_key:
            return Response({"error": "OpenAI API Key가 필요합니다. 설정 메뉴에서 입력해주세요."}, status=status.HTTP_400_BAD_REQUEST)
        
        # Prepare context (limiting text content for GPT context window while keeping relevant info)
        text_content = f"Title: {title}\nAbstract: {abstract}\n\nFull Text Snippet:\n{full_text[:15000]}"
        
        prompt = f"""
        Extract PICO (Population, Intervention, Comparison, Outcome) and Study Design information from the following medical literature text.
        Return the result in JSON format ONLY, matching the following structure strictly:
        {{
          "population": {{ "sample_size": number or null, "diagnosis": "string", "age_range": "string", "extracted_from": "specific sentence" }},
          "intervention": {{ "name": "string", "frequency": "string", "duration": "string", "extracted_from": "specific sentence" }},
          "comparison": {{ "type": "string", "extracted_from": "specific sentence" }},
          "outcome": {{ "primary_outcome": "string", "measurement_scales": ["string"], "extracted_from": "specific sentence" }},
          "study_design": {{ "blinding": "string", "allocation": "string", "design_type": "string" }},
          "statistical_summary": {{ "p_values": ["string"], "confidence_intervals": ["string"], "effect_sizes": ["string"] }},
          "extraction_confidence": number (0-1),
          "raw_evidence": ["list of key sentences supporting the extraction"]
        }}
        
        Rules:
        1. If information is missing, use null or empty string/array as appropriate.
        2. Ensure 'extracted_from' contains the exact sentence from the text if possible.
        3. 'raw_evidence' should be a list of 5-10 key sentences that provide context.
        
        Text to analyze:
        {text_content}
        """

        try:
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            }
            data = {
                "model": "gpt-4o",
                "messages": [
                    {
                        "role": "system", 
                        "content": "You are a highly skilled medical informatics expert specialized in systematic reviews and PICO data extraction for medical literature."
                    },
                    {"role": "user", "content": prompt}
                ],
                "response_format": { "type": "json_object" },
                "temperature": 0.1
            }

            response = http_requests.post(
                "https://api.openai.com/v1/chat/completions",
                headers=headers,
                json=data,
                timeout=60
            )
            
            if response.status_code != 200:
                error_info = response.json().get("error", {}).get("message", "Unknown error")
                return Response({"error": f"OpenAI API 오류: {error_info}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            pico_json_str = response.json()["choices"][0]["message"]["content"]
            pico_data = json.loads(pico_json_str)

            # Update DB if record_id provided
            if record_id:
                try:
                    rec = LiteratureRecord.objects.get(id=record_id)
                    rec.pico_data = pico_data
                    rec.pico_last_extracted_at = timezone.now() # Added
                    rec.save()
                    pico_data['record_id'] = record_id
                    pico_data['saved_to_db'] = True
                    pico_data['pico_last_extracted_at'] = rec.pico_last_extracted_at # Added
                except LiteratureRecord.DoesNotExist:
                    pico_data['saved_to_db'] = False

            return Response(pico_data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": f"ChatGPT 연동 오류: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class PicoConfirmView(APIView):
    """Finalize PICO extraction, moving to EXTRACTED status."""
    def post(self, request):
        record_id = request.data.get('record_id')
        pico_data = request.data.get('pico_data')
        try:
            rec = LiteratureRecord.objects.get(id=record_id)
            if pico_data:
                rec.pico_data = pico_data
            rec.status = LiteratureRecord.Status.EXTRACTED
            rec.pico_confirmed_at = timezone.now()
            rec.save()
            return Response({
                "ok": True, 
                "status": rec.status, 
                "pico_confirmed_at": rec.pico_confirmed_at
            }, status=status.HTTP_200_OK)
        except LiteratureRecord.DoesNotExist:
            return Response({"error": "Record not found"}, status=status.HTTP_404_NOT_FOUND)

class FulltextEligibleListView(APIView):
    """Return RCT_INCLUDED records waiting for full-text eligibility screening."""
    def get(self, request):
        project_id = request.query_params.get('project_id')
        qs = LiteratureRecord.objects.filter(
            status__in=[
                LiteratureRecord.Status.RCT_INCLUDED,
                LiteratureRecord.Status.FULLTEXT_INCLUDED,
                LiteratureRecord.Status.FULLTEXT_EXCLUDED,
            ]
        )
        if project_id:
            qs = qs.filter(project_id=project_id)
        records = list(qs.values('id', 'title', 'abstract', 'authors', 'year', 'pmid',
                                 'doi', 'source_db', 'status', 'exclusion_reason', 'reviewer_notes', 'full_text', 'pico_data'))
        return Response({"records": records, "count": len(records)}, status=status.HTTP_200_OK)

class FetchFullTextView(APIView):
    """Attempt to fetch full text from PMC for a given record."""
    def post(self, request):
        from .utils.fulltext_fetcher import FullTextFetcher
        record_id = request.data.get('record_id')
        try:
            rec = LiteratureRecord.objects.get(id=record_id)
            if not rec.pmid:
                return Response({"error": "No PMID available for this record"}, status=status.HTTP_400_BAD_REQUEST)
            
            fetcher = FullTextFetcher()
            text = fetcher.auto_fetch(rec.pmid)
            
            if text:
                rec.full_text = text
                rec.save()
                return Response({"ok": True, "full_text": text}, status=status.HTTP_200_OK)
            else:
                return Response({"error": "Full text not found or not Open Access in PMC"}, status=status.HTTP_404_NOT_FOUND)
        except LiteratureRecord.DoesNotExist:
            return Response({"error": "Record not found"}, status=status.HTTP_404_NOT_FOUND)

class UploadFullTextPDFView(APIView):
    """Handle PDF upload and extract text for a record."""
    def post(self, request):
        record_id = request.data.get('record_id')
        pdf_file = request.FILES.get('pdf_file')

        if not record_id or not pdf_file:
            return Response({"error": "Missing record_id or pdf_file"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            rec = LiteratureRecord.objects.get(id=record_id)
            
            # Extract text from PDF
            extracted_text = ""
            with pdfplumber.open(pdf_file) as pdf:
                for page in pdf.pages:
                    # Get page dimensions
                    width = page.width
                    height = page.height
                    
                    # Define regions for 2-column layout (Left side vs Right side)
                    left_bbox = (0, 0, width / 2, height)
                    right_bbox = (width / 2, 0, width, height)
                    
                    left_crop = page.within_bbox(left_bbox)
                    right_crop = page.within_bbox(right_bbox)
                    
                    # Extract text from left then right to preserve reading order
                    left_text = left_crop.extract_text()
                    right_text = right_crop.extract_text()
                    
                    if left_text:
                        extracted_text += left_text + "\n"
                    if right_text:
                        extracted_text += right_text + "\n"
                    
                    extracted_text += "\n" # Page break
            
            if not extracted_text.strip():
                return Response({"error": "Could not extract text from PDF"}, status=status.HTTP_422_UNPROCESSABLE_ENTITY)
            
            rec.full_text = extracted_text
            rec.save()
            return Response({"ok": True, "full_text": extracted_text}, status=status.HTTP_200_OK)

        except LiteratureRecord.DoesNotExist:
            return Response({"error": "Record not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class FulltextScreenView(APIView):
    """Proxy text to AI engine for eligibility screening. Does NOT save to DB."""
    def post(self, request):
        ai_url = os.getenv('AI_ENGINE_URL', 'http://ai_engine:8001')
        payload = {
            "title": request.data.get('title', ''),
            "abstract": request.data.get('abstract', ''),
            "full_text": request.data.get('full_text', ''),
            "criteria": request.data.get('criteria', []),
        }
        try:
            resp = http_requests.post(f"{ai_url}/api/v1/ai/screen_fulltext", json=payload, timeout=20)
            resp.raise_for_status()
            return Response(resp.json(), status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

class FulltextDecisionView(APIView):
    """Save researcher's final include/exclude decision + reason for full-text screening."""
    def post(self, request):
        record_id = request.data.get('record_id')
        decision = request.data.get('decision')  # 'include' | 'exclude'
        exclusion_reason = request.data.get('exclusion_reason', '')
        reviewer_notes = request.data.get('reviewer_notes', '')

        if not record_id or decision not in ('include', 'exclude'):
            return Response({"error": "Invalid parameters"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            rec = LiteratureRecord.objects.get(id=record_id)
            rec.status = (LiteratureRecord.Status.FULLTEXT_INCLUDED
                          if decision == 'include'
                          else LiteratureRecord.Status.FULLTEXT_EXCLUDED)
            rec.exclusion_reason = exclusion_reason if decision == 'exclude' else None
            rec.reviewer_notes = reviewer_notes
            # Also save manual full-text if provided
            if request.data.get('full_text'):
                rec.full_text = request.data.get('full_text')
            rec.save()
            return Response({"ok": True, "id": record_id, "status": rec.status}, status=status.HTTP_200_OK)
        except LiteratureRecord.DoesNotExist:
            return Response({"error": "Record not found"}, status=status.HTTP_404_NOT_FOUND)


class ProjectListView(APIView):
    """List all projects with per-project record statistics, or create a new project."""

    def get(self, request):
        projects = SearchProject.objects.all().order_by('-created_at')
        result = []
        for p in projects:
            qs = LiteratureRecord.objects.filter(project=p)
            total = qs.count()
            result.append({
                'id': p.id,
                'name': p.name,
                'description': p.description or '',
                'created_at': p.created_at.isoformat(),
                'updated_at': p.updated_at.isoformat(),
                'stats': {
                    'total': total,
                    'dedup_rejected': qs.filter(status=LiteratureRecord.Status.DEDUP_REJECTED).count(),
                    'rct_included': qs.filter(status__in=[
                        LiteratureRecord.Status.RCT_INCLUDED,
                        LiteratureRecord.Status.FULLTEXT_INCLUDED,
                        LiteratureRecord.Status.FULLTEXT_EXCLUDED,
                        LiteratureRecord.Status.EXTRACTED,
                        LiteratureRecord.Status.ROB_COMPLETED,
                    ]).count(),
                    'extracted': qs.filter(status__in=[
                        LiteratureRecord.Status.EXTRACTED,
                        LiteratureRecord.Status.ROB_COMPLETED,
                    ]).count(),
                }
            })
        return Response({'projects': result}, status=status.HTTP_200_OK)

    def post(self, request):
        name = request.data.get('name', '').strip()
        description = request.data.get('description', '').strip()
        if not name:
            return Response({"error": "Project name is required"}, status=status.HTTP_400_BAD_REQUEST)
        if SearchProject.objects.filter(name=name).exists():
            return Response({"error": "A project with this name already exists"}, status=status.HTTP_400_BAD_REQUEST)
        project = SearchProject.objects.create(name=name, description=description)
        return Response({
            'id': project.id,
            'name': project.name,
            'description': project.description,
            'created_at': project.created_at.isoformat(),
            'stats': {'total': 0, 'dedup_rejected': 0, 'rct_included': 0, 'extracted': 0}
        }, status=status.HTTP_201_CREATED)


class ProjectDetailView(APIView):
    """Retrieve, update, or delete a single project."""

    def get(self, request, pk):
        try:
            p = SearchProject.objects.get(pk=pk)
        except SearchProject.DoesNotExist:
            return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        qs = LiteratureRecord.objects.filter(project=p)
        return Response({
            'id': p.id, 'name': p.name, 'description': p.description or '',
            'created_at': p.created_at.isoformat(),
            'stats': {
                'total': qs.count(),
                'dedup_rejected': qs.filter(status=LiteratureRecord.Status.DEDUP_REJECTED).count(),
                'rct_included': qs.filter(status__in=[
                    LiteratureRecord.Status.RCT_INCLUDED,
                    LiteratureRecord.Status.FULLTEXT_INCLUDED,
                    LiteratureRecord.Status.FULLTEXT_EXCLUDED,
                    LiteratureRecord.Status.EXTRACTED,
                ]).count(),
                'extracted': qs.filter(status=LiteratureRecord.Status.EXTRACTED).count(),
            }
        })

    def patch(self, request, pk):
        try:
            p = SearchProject.objects.get(pk=pk)
        except SearchProject.DoesNotExist:
            return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        if 'name' in request.data:
            p.name = request.data['name'].strip() or p.name
        if 'description' in request.data:
            p.description = request.data['description']
        p.save()
        return Response({'id': p.id, 'name': p.name, 'description': p.description})

    def delete(self, request, pk):
        try:
            p = SearchProject.objects.get(pk=pk)
        except SearchProject.DoesNotExist:
            return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        # Prevent deleting the only remaining project
        if SearchProject.objects.count() <= 1:
            return Response({"error": "Cannot delete the last remaining project"}, status=status.HTTP_400_BAD_REQUEST)
        name = p.name
        p.delete()
        return Response({"ok": True, "deleted": name}, status=status.HTTP_200_OK)

class ExportRctExcelView(APIView):
    """Generate Excel file for RCT included records including PICO data."""
    def get(self, request):
        project_id = request.query_params.get('project_id')
        qs = LiteratureRecord.objects.filter(
            status__in=[
                LiteratureRecord.Status.RCT_INCLUDED,
                LiteratureRecord.Status.FULLTEXT_INCLUDED,
                LiteratureRecord.Status.EXTRACTED
            ]
        )
        if project_id:
            qs = qs.filter(project_id=project_id)
        
        # Create Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RCT_Included_Records"

        # Headers
        headers = [
            "ID", "Source DB", "Year", "Journal", "Authors", "Title", "PMID", "DOI", "Status",
            "P-Population (Diagnosis)", "P-Sample Size", "P-Age Range",
            "I-Intervention (Name)", "I-Frequency", "I-Duration",
            "C-Comparison", "O-Primary Outcome", "O-Scales",
            "Design", "Blinding", "Allocation",
            "P-values", "Effect Sizes", "Confidence"
        ]
        
        # Styling for header
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data Rows
        for row_num, rec in enumerate(qs, 2):
            pico = rec.pico_data or {}
            
            # Extract nested PICO values safely
            pop = pico.get('population', {})
            intv = pico.get('intervention', {})
            comp = pico.get('comparison', {})
            out = pico.get('outcome', {})
            design = pico.get('study_design', {})
            stats = pico.get('statistical_summary', {})
            
            data = [
                rec.id, rec.source_db, rec.year, rec.journal, rec.authors, rec.title, rec.pmid, rec.doi, rec.status,
                pop.get('diagnosis'), pop.get('sample_size'), pop.get('age_range'),
                intv.get('name'), intv.get('frequency'), intv.get('duration'),
                comp.get('type'), out.get('primary_outcome'), ", ".join(out.get('measurement_scales', [])),
                design.get('design_type'), design.get('blinding'), design.get('allocation'),
                ", ".join(stats.get('p_values', [])), ", ".join(stats.get('effect_sizes', [])),
                pico.get('extraction_confidence')
            ]
            
            for col_num, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else "")

        # Column Adjustments
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename=RCT_Extraction_List.xlsx'
        wb.save(response)
        return response

class ExportRctScreeningExcelView(APIView):
    """Excel export for RCT Screening page."""
    def get(self, request):
        project_id = request.query_params.get('project_id')
        qs = LiteratureRecord.objects.filter(status=LiteratureRecord.Status.SCREENING_PENDING)
        if project_id:
            qs = qs.filter(project_id=project_id)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RCT_Screening_Pending"
        
        headers = ["ID", "Source DB", "Year", "Journal", "Authors", "Title", "PMID", "DOI", "Status"]
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_num, rec in enumerate(qs, 2):
            data = [rec.id, rec.source_db, rec.year, rec.journal, rec.authors, rec.title, rec.pmid, rec.doi, rec.status]
            for col_num, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else "")
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = min(max_length + 2, 60)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=RCT_Screening_List.xlsx'
        wb.save(response)
        return response

class ExportFulltextScreeningExcelView(APIView):
    """Excel export for Full-text Screening page."""
    def get(self, request):
        project_id = request.query_params.get('project_id')
        qs = LiteratureRecord.objects.filter(
            status__in=[
                LiteratureRecord.Status.RCT_INCLUDED,
                LiteratureRecord.Status.FULLTEXT_INCLUDED,
                LiteratureRecord.Status.FULLTEXT_EXCLUDED
            ]
        )
        if project_id:
            qs = qs.filter(project_id=project_id)
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fulltext_Screening"
        
        headers = ["ID", "Source DB", "Year", "Journal", "Authors", "Title", "PMID", "DOI", "Status", "Exclusion Reason", "Reviewer Notes"]
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_num, rec in enumerate(qs, 2):
            data = [
                rec.id, rec.source_db, rec.year, rec.journal, rec.authors, rec.title, rec.pmid, rec.doi, rec.status,
                rec.exclusion_reason or "", rec.reviewer_notes or ""
            ]
            for col_num, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else "")

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = min(max_length + 2, 60)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Fulltext_Screening_List.xlsx'
        wb.save(response)
        return response

class RobAssessmentListView(APIView):
    """Return records that have completed PICO extraction, ready for ROB assessment."""
    def get(self, request):
        project_id = request.query_params.get('project_id')
        qs = LiteratureRecord.objects.filter(
            status__in=[
                LiteratureRecord.Status.EXTRACTED,
                LiteratureRecord.Status.ROB_COMPLETED
            ]
        )
        if project_id:
            qs = qs.filter(project_id=project_id)
        records = list(qs.values('id', 'title', 'abstract', 'authors', 'year', 'pmid', 'doi', 'source_db', 'status', 'full_text', 'pico_data', 'rob_data', 'rob_last_saved_at', 'rob_completed_at'))
        return Response({"records": records, "count": len(records)}, status=status.HTTP_200_OK)

class RobSaveView(APIView):
    """Save ROB assessment data for a record."""
    def post(self, request):
        print(f"DEBUG: RobSaveView - data: {request.data}")
        record_id = request.data.get('record_id')
        rob_data = request.data.get('rob_data')
        
        # Explicit boolean parsing
        complete_raw = request.data.get('complete', True)
        if isinstance(complete_raw, str):
            complete = complete_raw.lower() == 'true'
        else:
            complete = bool(complete_raw)
            
        print(f"DEBUG: RobSaveView - record_id: {record_id}, complete: {complete}")
        
        if not record_id or rob_data is None:
            return Response({"error": "Missing record_id or rob_data"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            rec = LiteratureRecord.objects.get(id=record_id)
            rec.rob_data = rob_data
            now = timezone.now()
            rec.rob_last_saved_at = now
            if complete:
                rec.status = LiteratureRecord.Status.ROB_COMPLETED
                rec.rob_completed_at = now
            else:
                # Ensure status stays as EXTRACTED if it's a temp save
                if rec.status == LiteratureRecord.Status.ROB_COMPLETED and not complete:
                     # If it was completed but user clicked temp save, maybe keep it completed or revert?
                     # Revert to EXTRACTED to move back to pending tab if requested
                     rec.status = LiteratureRecord.Status.EXTRACTED
                elif rec.status != LiteratureRecord.Status.ROB_COMPLETED:
                     rec.status = LiteratureRecord.Status.EXTRACTED

            rec.save()
            return Response({
                "ok": True, 
                "id": record_id, 
                "rob_last_saved_at": rec.rob_last_saved_at.isoformat() if rec.rob_last_saved_at else None, 
                "rob_completed_at": rec.rob_completed_at.isoformat() if rec.rob_completed_at else None,
                "status": rec.status
            }, status=status.HTTP_200_OK)
        except LiteratureRecord.DoesNotExist:
            return Response({"error": "Record not found"}, status=status.HTTP_404_NOT_FOUND)

class RobPredictView(APIView):
    """Proxy to AI engine for automated ROB suggestions."""
    def post(self, request):
        ai_url = os.getenv('AI_ENGINE_URL', 'http://ai_engine:8001')
        payload = {
            "title": request.data.get('title', ''),
            "abstract": request.data.get('abstract', ''),
            "full_text": request.data.get('full_text', ''),
        }
        try:
            resp = http_requests.post(f"{ai_url}/api/v1/ai/predict_rob", json=payload, timeout=25)
            resp.raise_for_status()
            return Response(resp.json(), status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

class RobPredictGptView(APIView):
    """Enhanced ROB assessment using GPT-4o and user API key."""
    def post(self, request):
        api_key = request.data.get('api_key')
        title = request.data.get('title', '')
        abstract = request.data.get('abstract', '')
        full_text = request.data.get('full_text', '')

        if not api_key:
            return Response({"error": "OpenAI API Key가 필요합니다. 설정 메뉴에서 입력해주세요."}, status=status.HTTP_400_BAD_REQUEST)
        
        # Limit text for context window
        text_content = f"Title: {title}\nAbstract: {abstract}\n\nFull Text Snippet:\n{full_text[:18000]}"
        
        prompt = f"""
        Analyze the following research article for Risk of Bias using Cochrane ROB 1.0 (7 domains).
        For each domain, provide:
        1. A decision: 'low', 'high', or 'unclear'.
        2. A short supporting sentence (evidence) directly from the text if possible.
        3. A confidence score (0-1).

        Domains:
        - d1: Random sequence generation (Selection bias)
        - d2: Allocation concealment (Selection bias)
        - d3: Blinding of participants and personnel (Performance bias)
        - d4: Blinding of outcome assessment (Detection bias)
        - d5: Incomplete outcome data (Attrition bias)
        - d6: Selective reporting (Reporting bias)
        - d7: Other bias (Any other potential source of bias)

        Return the result in JSON format ONLY, matching the structure:
        {{
          "domains": {{
            "d1": {{ "decision": "string", "evidence": "string", "confidence": number }},
            "d2": {{ ... }},
            ...
            "d7": {{ ... }}
          }}
        }}
        
        Text to analyze:
        {text_content}
        """

        try:
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            }
            data = {
                "model": "gpt-4o",
                "messages": [
                    {
                        "role": "system", 
                        "content": "You are a professional medical literature reviewer specialized in Cochrane ROB assessment."
                    },
                    {"role": "user", "content": prompt}
                ],
                "response_format": { "type": "json_object" },
                "temperature": 0.1
            }

            response = http_requests.post(
                "https://api.openai.com/v1/chat/completions",
                headers=headers,
                json=data,
                timeout=90
            )
            
            if response.status_code != 200:
                err_msg = response.json().get("error", {}).get("message", "Unknown error")
                return Response({"error": f"OpenAI API 오류: {err_msg}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            rob_data = json.loads(response.json()["choices"][0]["message"]["content"])
            return Response(rob_data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ExportRobExcelView(APIView):
    """Excel export for ROB Assessment results."""
    def get(self, request):
        project_id = request.query_params.get('project_id')
        qs = LiteratureRecord.objects.filter(
            status__in=[
                LiteratureRecord.Status.EXTRACTED,
                LiteratureRecord.Status.ROB_COMPLETED
            ]
        )
        if project_id:
            qs = qs.filter(project_id=project_id)
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ROB_Assessment"
        
        # Base headers
        headers = ["ID", "Source DB", "Year", "Journal", "Authors", "Title", "PMID", "DOI", "Status"]
        
        # ROB Domain headers (7 domains)
        domains = [
            ('d1', 'Random sequence generation'),
            ('d2', 'Allocation concealment'),
            ('d3', 'Blinding of participants and personnel'),
            ('d4', 'Blinding of outcome assessment'),
            ('d5', 'Incomplete outcome data'),
            ('d6', 'Selective reporting'),
            ('d7', 'Other bias'),
        ]
        
        for _, name in domains:
            headers.append(f"{name} (Risk)")
            headers.append(f"{name} (Comment)")
            
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_num, rec in enumerate(qs.order_by('id'), 2):
            row_data = [
                rec.id, rec.source_db, rec.year, rec.journal, rec.authors, rec.title, rec.pmid, rec.doi, rec.status
            ]
            
            rob = rec.rob_data or {}
            for dom_id, _ in domains:
                d_val = rob.get(dom_id, {})
                row_data.append(d_val.get('decision', ''))
                row_data.append(d_val.get('comment', ''))
            
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else "")

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=ROB_Assessment_Results.xlsx'
        wb.save(response)
        return response
